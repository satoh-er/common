import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from openpyxl.utils import get_column_letter

def insert_images_into_excel(folder_path, output_file):
    wb = Workbook()
    ws = wb.active
    
    files = [f for f in sorted(os.listdir(folder_path)) if f.endswith('.png')]
    files = sorted(files, key=lambda x: int(x.split('_')[1]))  # 'case_x_' の x 部分でソート
    max_width = 500  # 最大の画像幅
    max_height = 500  # 最大の画像高さ
    
    col = 1
    row = 1
    images_per_row = 8  # 1行に配置する画像の数

    for i, file in enumerate(files):
        print(file)
        img_path = os.path.join(folder_path, file)
        
        pil_img = PILImage.open(img_path)
        ratio = min(max_width / pil_img.width, max_height / pil_img.height)
        new_width = int(pil_img.width * ratio)
        new_height = int(pil_img.height * ratio)
        pil_img = pil_img.resize((new_width, new_height), PILImage.LANCZOS)
        
        # 一時ファイルに一意の名前を使用
        temp_img_path = f'temp_{i}.png'
        pil_img.save(temp_img_path)
        
        img = Image(temp_img_path)
        img.anchor = ws.cell(row=row, column=col).coordinate
        ws.add_image(img)
        
        # ファイル名の位置を調整（画像の直下に配置）
        ws.cell(row=row + 1, column=col).value = file
        
        # 列の幅と行の高さを適切に調整
        ws.row_dimensions[row].height = new_height * 0.75
        ws.column_dimensions[get_column_letter(col)].width = new_width * 0.14
        
        col += 1
        if col > images_per_row:
            col = 1
            row += 2  # 画像とファイル名の間隔を1行に設定

        if row > 24 * 2:  # 24行分の画像を超えた場合は終了
            break

    wb.save(output_file)

    # 保存後、使用した一時ファイルを削除
    for i in range(len(files)):
        
        os.remove(f'temp_{i}.png')

# 使用例
folder_path = 'fig'
output_file = 'output.xlsx'
insert_images_into_excel(folder_path, output_file)
